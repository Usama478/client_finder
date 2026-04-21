from fastapi import APIRouter, Depends, HTTPException, Body, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
import io
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.db.session import get_db
from app.models.search_result import SearchResult
from app.core.security import get_current_user
from app.services.credit_service import check_credits, deduct_credits
from app.services.activity_service import log_activity

router = APIRouter(prefix="/api/v1/export", tags=["export"])

COLUMN_HEADERS = [
    "Business Name",
    "Address",
    "Website",
    "Phone Number",
    "Emails",
    "Social Links",
    "Relevance Score",
    "Relevance Reason",
    "Verification Score",
    "Verification Reason",
    "Risk Flags",
    "Verification Status",
    "Trust: Website Live",
    "Trust: SSL Certificate",
    "Trust: Domain Age",
    "Trust: Privacy Policy",
    "Trust: About Page",
    "Trust: Contact Page",
    "Trust: Social Media",
    "Trust: Email Validity",
    "Trust: Legal Registration",
    "Trust: Risk Flags",
    "Trust: Legitimacy Score",
    "Trust: Contactability",
    "Trust: Wholesale Page",
    "Trust: LinkedIn",
    "Trust: Employee Range",
    "Trust: Revenue Band",
]


def _build_row(client) -> list:
    social_links = client.social_links or {}
    if social_links:
        socials_str = ", ".join(f"{k}: {v}" for k, v in social_links.items() if v)
    else:
        socials_str = ""

    risk_str = ", ".join(client.risk_flags) if getattr(client, "risk_flags", None) else ""

    row = [
        client.business_name or "",
        client.address or "",
        client.website or "",
        client.phone_number or "",
        client.email_found or "",
        socials_str,
        client.relevance_score if client.relevance_score is not None else "Pending",
        client.relevance_reason or "",
        client.verification_score if client.verification_score is not None else "Pending",
        client.verification_reason or "",
        risk_str,
        client.verification_status or "",
    ]

    # Trust columns - read from actual SearchResult fields
    artifacts = client.verification_artifacts or {}
    
    # Trust: Website Live
    website_live = "Yes" if client.verification_status != "pending" else "No"
    row.append(website_live)
    
    # Trust: SSL Certificate
    ssl_valid = artifacts.get("ssl_valid", "-") or "-"
    row.append(str(ssl_valid))
    
    # Trust: Domain Age
    domain_age = f"{client.domain_age_years} years" if client.domain_age_years is not None else "-"
    row.append(domain_age)
    
    # Trust: Privacy Policy
    if client.has_policy_pages is None:
        privacy_policy = "-"
    elif client.has_policy_pages:
        privacy_policy = "Found"
    else:
        privacy_policy = "Not Found"
    row.append(privacy_policy)
    
    # Trust: About Page
    if client.has_about_page is None:
        about_page = "-"
    elif client.has_about_page:
        about_page = "Found"
    else:
        about_page = "Not Found"
    row.append(about_page)
    
    # Trust: Contact Page
    if client.has_contact_page is None:
        contact_page = "-"
    elif client.has_contact_page:
        contact_page = "Found"
    else:
        contact_page = "Not Found"
    row.append(contact_page)
    
    # Trust: Social Media
    social_links_data = client.social_links or {}
    if social_links_data:
        social_count = len([k for k, v in social_links_data.items() if v])
        social_media = f"{social_count} profiles" if social_count > 0 else "-"
    else:
        social_media = "-"
    row.append(social_media)
    
    # Trust: Email Validity
    email_validity = client.email_type if client.email_type is not None else "-"
    row.append(email_validity)
    
    # Trust: Legal Registration
    legal_registration = artifacts.get("legal_registration", "-") or "-"
    row.append(str(legal_registration))
    
    # Trust: Risk Flags
    risk_flags_data = client.risk_flags or []
    risk_flags_str = ", ".join(risk_flags_data) if risk_flags_data else "None"
    row.append(risk_flags_str)
    
    # Trust: Legitimacy Score
    legitimacy = str(client.legitimacy_score) if client.legitimacy_score is not None else "-"
    row.append(legitimacy)
    
    # Trust: Contactability
    contactability = str(client.contactability_score) if client.contactability_score is not None else "-"
    row.append(contactability)
    
    # Trust: Wholesale Page
    if client.wholesale_page_url is not None:
        wholesale = client.wholesale_page_url
    elif client.wholesale_page_found:
        wholesale = "Found"
    else:
        wholesale = "Not Found"
    row.append(wholesale)
    
    # Trust: LinkedIn
    linkedin = client.linkedin_company_url if client.linkedin_company_url is not None else "-"
    row.append(linkedin)
    
    # Trust: Employee Range
    employee_range = client.employee_range if client.employee_range is not None else "-"
    row.append(employee_range)
    
    # Trust: Revenue Band
    revenue_band = client.revenue_band if client.revenue_band is not None else "-"
    row.append(revenue_band)

    return row


@router.post("")
def export_clients(
    result_ids: Optional[List[str]] = Body(None),
    status: Optional[str] = Query(None),
    format: str = Query("excel"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Exports clients to Excel or CSV.

    Priority:
      1. result_ids body → export only those rows
      2. status query param → filter by verification_status
      3. Neither → export all saved clients for the authenticated user
    """
    try:
        query = db.query(SearchResult).filter(
            SearchResult.is_saved_client == True,
            SearchResult.user_id == current_user.user_id,
        )

        if result_ids and len(result_ids) > 0:
            numeric_ids = [int(x) for x in result_ids if str(x).isdigit()]
            query = query.filter(
                SearchResult.place_id.in_(result_ids)
                | SearchResult.result_id.in_(numeric_ids)
            )
        elif status:
            query = query.filter(SearchResult.verification_status == status)

        check_credits(db, current_user.user_id, 5)

        clients = query.all()

        if not clients:
            raise HTTPException(status_code=404, detail="No clients found to export.")

        deduct_credits(db, current_user.user_id, 5, "export", reference_type="export")
        db.commit()
        try:
            log_activity(db, current_user.user_id, "export", metadata={"format": format, "count": len(clients)}, credits_consumed=5)
            db.commit()
        except Exception:
            pass

        timestamp = datetime.now().strftime("%Y-%m-%d")

        # ── CSV branch ────────────────────────────────────────────────────────
        if format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(COLUMN_HEADERS)
            for client in clients:
                writer.writerow(_build_row(client))

            output.seek(0)
            filename = f"Client_List_{timestamp}.csv"
            response_headers = {
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers=response_headers,
            )

        # ── Excel branch (default) ────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Client List"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F46E5", end_color="4F46E5", fill_type="solid"
        )

        for col_num, header_title in enumerate(COLUMN_HEADERS, 1):
            cell = ws.cell(row=1, column=col_num, value=header_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        for row_num, client in enumerate(clients, 2):
            row_data = _build_row(client)
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                # Relevance Reason (col 8) and Verification Reason (col 10) get wrap
                if col_num in [8, 10]:
                    cell.alignment = wrap_alignment
                else:
                    cell.alignment = Alignment(vertical="top")

        column_widths = {
            1: 30,   # Business Name
            2: 40,   # Address
            3: 30,   # Website
            4: 20,   # Phone
            5: 35,   # Emails
            6: 35,   # Socials
            7: 15,   # Relevance Score
            8: 60,   # Relevance Reason
            9: 15,   # Verification Score
            10: 60,  # Verification Reason
            11: 30,  # Risk Flags
            12: 18,  # Verification Status
            13: 20,  # Trust: Website Live
            14: 20,  # Trust: SSL Certificate
            15: 20,  # Trust: Domain Age
            16: 20,  # Trust: Privacy Policy
            17: 20,  # Trust: About Page
            18: 20,  # Trust: Contact Page
            19: 20,  # Trust: Social Media
            20: 20,  # Trust: Email Validity
            21: 20,  # Trust: Legal Registration
            22: 20,  # Trust: Risk Flags
            23: 20,  # Trust: Legitimacy Score
            24: 20,  # Trust: Contactability
            25: 20,  # Trust: Wholesale Page
            26: 30,  # Trust: LinkedIn
            27: 20,  # Trust: Employee Range
            28: 20,  # Trust: Revenue Band
        }

        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        filename = f"Client_List_{timestamp}.xlsx"
        response_headers = {
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=response_headers,
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
